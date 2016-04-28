# -*- coding: utf-8 -*-

import openpyxl
import re
import csv
import os

WHITE_LIST = [
    u"公開WEB情報",
    u"履歴書的経歴",
    u"人物像",
    u"創造的経歴"
]

def get_excels():
    for root, dirs, files in os.walk('interview'):
        for f in files:
            if f.endswith('.xlsx'):
                yield root + '/' + f

def main():
    import unicodecsv as csv
    wr = csv.writer(file('t.csv', 'w'), encoding='utf-8')

    from collections import Counter
    excels = get_excels()
    buf = []
    for f in excels:
        infile = file(f)
        if '繝' in f:
            f = f.decode('utf-8', 'ignore').encode('sjis', 'ignore')
        print f

        wb = openpyxl.load_workbook(infile)
        wss = [ws for ws in wb.worksheets
               if not ws.title.startswith(u'サンプル')
               if not ws.title.startswith(u"未踏名鑑の作成に当たって")
               if not ws.title.startswith(u"進め方・注意事項")]
        if all(x in [y.title for y in wss] for x in WHITE_LIST) and len(wss) == 4:
            pass
        else:
            print 'ERROR', f

        #buf.append(len(wss))
        if 0:
            for ws in wss:
                print ws.title,
            print

        sheets = [None] * 4
        for ws in wss:
            for i in range(4):
                if ws.title == WHITE_LIST[i]:
                    sheets[i] = ws

        wr.writerow(["# {}".format(f)])
        # 公開WEB情報
        wr.writerow(["## 公開WEB情報"])
        s = sheets[0]
        wr.writerows([
            [c.value for c in row]
            for row in s.get_squared_range(3, 13, 5, s.max_row)])

        # 履歴書的経歴
        wr.writerow(["## 履歴書的経歴"])
        s = sheets[1]
        wr.writerows([
            [c.value for c in row]
            for row in s.get_squared_range(2, 14, 9, s.max_row)])

        # 人物像
        wr.writerow(["## 人物像"])
        s = sheets[2]
        wr.writerows([
            [c.value for c in row]
            for row in s.get_squared_range(2, 11, 4, s.max_row)])

        # 創造的経歴
        wr.writerow(["## 創造的経歴"])
        s = sheets[3]
        wr.writerows([
            [c.value for c in row]
            for row in s.get_squared_range(2, 15, 10, s.max_row)])

        #return s
    #print Counter(buf)


s = main()
